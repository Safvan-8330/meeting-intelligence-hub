from multiprocessing import context

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
import os, io, csv, textwrap, json
from datetime import datetime
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.database import supabase
from pydantic import BaseModel
from app.services.extractor import extract_full_intelligence
from app.auth import verify_user # 👈 IMPORT THE BOUNCER

# Resolve uploads directory relative to the backend root so routes work
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')

router = APIRouter()

def safe_split_text(text, width=85):
    if not text:
        return []
    # Strip non-ASCII to prevent PDF crashes
    clean = str(text).encode('ascii', 'ignore').decode('ascii')
    return textwrap.wrap(clean, width=width)

# 🔒 SECURED: Full Report Data
@router.get("/full-report/{filename}")
async def get_full_report(filename: str, user_id: str = Depends(verify_user)):
    # 👇 Added .eq("user_id", user_id) so users can only fetch their own meeting data
    res = supabase.table("meetings").select("id").eq("filename", filename).eq("user_id", user_id).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Meeting not found or access denied")
    m_id = res.data[0]['id']

    decisions = supabase.table("decisions").select("text").eq("meeting_id", m_id).execute()
    actions = supabase.table("action_items").select("*").eq("meeting_id", m_id).execute()
    sentiments = supabase.table("sentiments").select("*").eq("meeting_id", m_id).execute()
    users = supabase.table("meeting_users").select("user_name").eq("meeting_id", m_id).execute()
    trans_res = supabase.table("transcripts").select("raw_text").eq("meeting_id", m_id).execute()

    transcript = ""
    if trans_res and getattr(trans_res, 'data', None):
        if isinstance(trans_res.data, list):
            parts = [r.get('raw_text', '') for r in trans_res.data if r.get('raw_text')]
            transcript = "\n\n".join(parts)
        elif isinstance(trans_res.data, dict):
            transcript = trans_res.data.get('raw_text', '')

    decisions_list = decisions.data or []
    actions_list = actions.data or []
    sentiments_list = sentiments.data or []
    users_list = users.data or []

    return {
        "filename": filename,
        "transcript": transcript,
        "decisions": [d.get('text') for d in decisions_list],
        "action_items": actions_list,
        "participants": [u.get('user_name') for u in users_list],
        "speaker_sentiment": [s for s in sentiments_list if s.get('type') == 'speaker'],
        "segment_sentiment": [s for s in sentiments_list if s.get('type') == 'segment']
    }

# 🔒 SECURED: Excel Export
@router.get("/export/excel/{filename}")
async def export_excel(filename: str, user_id: str = Depends(verify_user)):
    try:
        data = await get_full_report(filename, user_id) # Pass user_id down
    except Exception:
        raise HTTPException(status_code=404, detail="Meeting data not found in database.")

    wb = Workbook()
    
    # --- SHEET 1: TASKS & DECISIONS ---
    ws1 = wb.active
    ws1.title = "Tasks and Decisions"
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(horizontal="center", vertical="top")

    headers1 = ["Item Type", "Assignee", "Description / Task", "Due Date"]
    ws1.append(headers1)

    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 60
    ws1.column_dimensions['D'].width = 15

    decisions = data.get('decisions', [])
    action_items = data.get('action_items', [])

    for d in decisions:
        ws1.append(["Decision", "-", d or "", "-"])

    for a in action_items:
        assignee = (a.get('assignee') or a.get('who') or '').strip()
        desc = (a.get('task') or a.get('what') or '').strip()
        due = (a.get('due_date') or a.get('by_when') or 'TBD')
        ws1.append(["Action Item", assignee, desc, due])

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=4):
        row[2].alignment = wrap_alignment
        row[0].alignment = center_alignment
        row[3].alignment = center_alignment

    # --- SHEET 2: PARTICIPANTS & SENTIMENT ---
    ws2 = wb.create_sheet(title="Sentiment Analysis")
    headers2 = ["Name/Participant", "Sentiment Label", "Score (-1 to 1)"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    speaker_sentiment = data.get('speaker_sentiment', [])
    for s in speaker_sentiment:
        score = s.get('score', 0)
        label = "Positive" if score >= 0.2 else "Negative" if score <= -0.2 else "Neutral"
        ws2.append([s.get('label') or s.get('name') or 'Unknown', label, score])

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    clean_filename = filename.replace(".txt", "")
    headers = {
        'Content-Disposition': f'attachment; filename="{clean_filename}_Report.xlsx"'
    }
    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

# 🔒 SECURED: PDF Export
@router.get("/export/pdf/{filename}")
async def export_pdf(filename: str, user_id: str = Depends(verify_user)):
    try:
        data = await get_full_report(filename, user_id) # Pass user_id down
    except Exception:
        raise HTTPException(status_code=404, detail="Meeting data not found in database.")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # --- HEADER BLOCK ---
    pdf.set_fill_color(30, 41, 59)
    pdf.rect(0, 0, 210, 45, 'F')
    
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", style="B", size=24)
    pdf.set_xy(15, 12)
    pdf.cell(0, 10, "Meeting Intelligence Brief", ln=True)
    
    pdf.set_font("Helvetica", size=10)
    pdf.set_xy(15, 24)
    pdf.cell(0, 10, f"Source File: {filename}", ln=True)
    pdf.set_xy(15, 30)
    pdf.cell(0, 10, f"Generated: {datetime.utcnow().strftime('%B %d, %Y at %H:%M UTC')}", ln=True)

    # --- RESET COLORS ---
    pdf.set_y(55)
    pdf.set_text_color(30, 41, 59)

    def draw_section_header(title):
        pdf.ln(5)
        pdf.set_font("Helvetica", style="B", size=14)
        pdf.set_text_color(79, 70, 229)
        pdf.cell(0, 10, title, ln=True)
        pdf.set_draw_color(226, 232, 240)
        pdf.line(15, pdf.get_y(), 195, pdf.get_y())
        pdf.ln(4)
        pdf.set_text_color(51, 65, 85)
        pdf.set_font("Helvetica", size=11)

    participants = data.get('participants', [])
    draw_section_header("Meeting Participants")
    if participants:
        pdf.multi_cell(0, 6, " * " + "\n * ".join(participants))
    else:
        pdf.cell(0, 6, "No participants identified.", ln=True)

    draw_section_header("Sentiment Analysis")
    speaker_sent = data.get('speaker_sentiment', [])
    if speaker_sent:
        for s in speaker_sent:
            name = s.get('label') or s.get('name') or 'Unknown'
            score = s.get('score', 0)
            vibe = "Positive" if score >= 0.1 else "Negative" if score <= -0.1 else "Neutral"
            score_txt = f"{score:.2f}" if isinstance(score, (int, float)) else "N/A"
            pdf.cell(0, 7, f"* {name}: {vibe} (Score: {score_txt})", ln=True)
    else:
        pdf.cell(0, 7, "No sentiment data available.", ln=True)

    decisions = data.get('decisions', [])
    draw_section_header("Key Decisions")
    if decisions:
        for d in decisions:
            pdf.set_x(15)
            for line in safe_split_text(f"+  {d}", width=95):
                pdf.cell(0, 6, line, ln=True)
            pdf.ln(2)
    else:
        pdf.cell(0, 6, "No formal decisions recorded.", ln=True)

    action_items = data.get('action_items', [])
    draw_section_header("Action Items Tracker")
    
    if action_items:
        for a in action_items:
            assignee = a.get('assignee') or a.get('who') or 'Unassigned'
            task = a.get('task') or a.get('what') or 'No description'
            due = a.get('due_date') or a.get('by_when') or 'TBD'

            pdf.set_font("Helvetica", style="B", size=11)
            pdf.set_text_color(15, 23, 42)
            pdf.cell(25, 6, f"Assignee: ", ln=False)
            pdf.set_font("Helvetica", size=11)
            pdf.cell(0, 6, assignee, ln=True)
            
            pdf.set_font("Helvetica", style="B", size=11)
            pdf.cell(25, 6, f"Due Date: ", ln=False)
            pdf.set_font("Helvetica", size=11)
            pdf.set_text_color(225, 29, 72)
            pdf.cell(0, 6, due, ln=True)
            
            pdf.set_text_color(51, 65, 85)
            for line in safe_split_text(f"Task: {task}", width=90):
                pdf.cell(0, 6, line, ln=True)
            pdf.ln(4)
    else:
        pdf.cell(0, 6, "No action items recorded.", ln=True)

    pdf.add_page()
    draw_section_header("Full Meeting Transcript")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 116, 139)
    transcript = data.get('transcript', 'No transcript available.')
    pdf.multi_cell(0, 5, transcript)

    return StreamingResponse(io.BytesIO(pdf.output()), media_type="application/pdf")

# 🔒 SECURED: Fetch Single Analysis
@router.get("/{filename}")
async def get_analysis(filename: str, user_id: str = Depends(verify_user)):
    # 👇 Security check before touching the cache!
    res = supabase.table("meetings").select("id").eq("filename", filename).eq("user_id", user_id).execute()
    if not res.data:
        raise HTTPException(status_code=403, detail="Access denied or meeting not found.")

    file_path = os.path.join(UPLOAD_DIR, filename)
    cache_path = os.path.join(UPLOAD_DIR, f"{filename}_analysis.json")

    if os.path.exists(cache_path):
        with open(cache_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return {"analysis": data}

    try:
        data = await get_full_report(filename, user_id)
        analysis = {
            "transcript": data.get("transcript", ""),
            "decisions": data.get("decisions", []),
            "action_items": data.get("action_items", []),
            "participants": data.get("participants", []),
            "speaker_sentiment": data.get("speaker_sentiment", []),
            "segment_sentiment": data.get("segment_sentiment", []),
        }
        return {"analysis": analysis}
    except HTTPException:
        if os.path.exists(file_path):
            with open(file_path, 'r', encoding='utf-8') as f:
                txt = f.read()
            return {"analysis": {"transcript": txt, "decisions": [], "action_items": [], "participants": []}}
        raise

# 🔒 SECURED: Manual Sync
@router.post('/sync/{filename}')
async def sync_analysis(filename: str, user_id: str = Depends(verify_user)):
    file_path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail='Transcript file not found')

    try:
        # 👇 Tell Supabase this file belongs to user_id during upsert
        supabase.table('meetings').upsert({'filename': filename, 'user_id': user_id}, on_conflict='filename').execute()
    except Exception as e:
        print(f"⚠️ Could not upsert meeting: {e}")

    res = supabase.table('meetings').select('id').eq('filename', filename).execute()
    if not res or not getattr(res, 'data', None):
        raise HTTPException(status_code=500, detail='Failed to resolve meeting id')
    m_id = res.data[0]['id']

    ai_data, raw_text = extract_full_intelligence(file_path)
    if not ai_data:
        raise HTTPException(status_code=503, detail='AI analysis failed; try again later')

    analysis = {
        'transcript': raw_text,
        'decisions': ai_data.get('decisions', []),
        'action_items': ai_data.get('action_items', []),
        'participants': ai_data.get('users', []),
        'speaker_sentiment': ai_data.get('speaker_sentiment', []),
        'segment_sentiment': ai_data.get('segment_sentiment', [])
    }
    cache_path = os.path.join(UPLOAD_DIR, f"{filename}_analysis.json")
    try:
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(analysis, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ Failed to write analysis cache: {e}")

    for table in ('decisions', 'action_items', 'sentiments', 'meeting_users'):
        try:
            supabase.table(table).delete().eq('meeting_id', m_id).execute()
        except Exception:
            pass

    try:
        supabase.table('transcripts').upsert({'meeting_id': m_id, 'raw_text': raw_text}).execute()

        if analysis['decisions']:
            dec_list = [{'meeting_id': m_id, 'text': d} for d in analysis['decisions']]
            supabase.table('decisions').insert(dec_list).execute()

        act_items = ai_data.get('action_items', [])
        normalized = []
        for a in act_items:
            who = a.get('who') or a.get('assignee') or a.get('owner') or None
            what = a.get('what') or a.get('task') or a.get('title') or None
            when = a.get('by_when') or a.get('due_date') or a.get('deadline') or None
            normalized.append({'meeting_id': m_id, 'assignee': who, 'task': what, 'due_date': when})
        if normalized:
            supabase.table('action_items').insert(normalized).execute()

        sent_list = []
        for s in ai_data.get('speaker_sentiment', []):
            sent_list.append({'meeting_id': m_id, 'type': 'speaker', 'label': s.get('name'), 'score': s.get('score')})
        for seg in ai_data.get('segment_sentiment', []):
            sent_list.append({'meeting_id': m_id, 'type': 'segment', 'label': seg.get('timestamp'), 'score': seg.get('score')})
        if sent_list:
            supabase.table('sentiments').insert(sent_list).execute()

        users = ai_data.get('users') or analysis.get('participants') or []
        if users:
            user_list = [{'meeting_id': m_id, 'user_name': u} for u in users]
            supabase.table('meeting_users').insert(user_list).execute()

    except Exception as e:
        print(f"❌ Error syncing analysis to Supabase: {e}")
        raise HTTPException(status_code=500, detail='Failed to sync analysis to database')

    return {'result': 'ok', 'analysis': analysis}

class ChatQuery(BaseModel):
    query: str  # 👈 Change this to 'query' to match the frontend JS
    filename: str = "all"

# 🔒 SECURED: God Mode Global Search
@router.post("/global-query")
async def global_search(query: ChatQuery, user_id: str = Depends(verify_user)):
    try:
        # 1. Only get IDs for meetings OWNED by this user
        meetings_res = supabase.table("meetings").select("id, filename").eq("user_id", user_id).execute()
        
        if not meetings_res.data:
            return {"answer": "I couldn't find any meetings in your account to search through."}
        
        # Create a list of valid IDs to filter the other tables
        meeting_ids = [m['id'] for m in meetings_res.data]
        meetings_map = {m['id']: m['filename'] for m in meetings_res.data}
        
        # 2. Setup context object
        context_data = {m_id: {"filename": fname, "transcript": "", "actions": [], "decisions": []} 
                        for m_id, fname in meetings_map.items()}

        # 3. SECURE FETCH: Only pull data belonging to the filtered meeting_ids
        # This prevents pulling the entire database into memory
        trans_res = supabase.table("transcripts").select("meeting_id, raw_text").in_("meeting_id", meeting_ids).execute()
        actions_res = supabase.table("action_items").select("meeting_id, assignee, task, due_date").in_("meeting_id", meeting_ids).execute()
        decisions_res = supabase.table("decisions").select("meeting_id, text").in_("meeting_id", meeting_ids).execute()

        # 4. Map the data to the correct meetings
        for t in (trans_res.data or []):
            context_data[t['meeting_id']]['transcript'] += t.get('raw_text', '') + "\n"
        
        for a in (actions_res.data or []):
            assignee = a.get('assignee') or 'Unassigned'
            task = a.get('task') or ''
            due = a.get('due_date') or 'TBD'
            context_data[a['meeting_id']]['actions'].append(f"{assignee} needs to: {task} (Due: {due})")

        for d in (decisions_res.data or []):
            context_data[d['meeting_id']]['decisions'].append(d.get('text', ''))

        # 5. Build the text context for the AI
        context = ""
        for m_id, data in context_data.items():
            context += f"\n========================\n"
            context += f"MEETING: {data['filename']}\n"
            context += f"========================\n"
            
            if data['decisions']:
                context += "► KEY DECISIONS MADE:\n" + "\n".join([f"  - {d}" for d in data['decisions']]) + "\n\n"
            
            if data['actions']:
                context += "► ACTION ITEMS ASSIGNED:\n" + "\n".join([f"  - {a}" for a in data['actions']]) + "\n\n"
                
            context += "► RAW TRANSCRIPT:\n" + data['transcript'] + "\n"

        system_prompt = (
            "You are an elite Meeting Intelligence Assistant. Use the provided Action Items, "
            "Decisions, and transcripts to answer accurately. Mention filenames when referring to specific meetings."
        )
        
        # Note: Ensure your ChatQuery model has a 'query' or 'question' field that matches your frontend!
        # Change query.question to query.query
        user_prompt = f"Context from all meetings:\n{context}\n\nUser Question: {query.query}"

        from app.services.extractor import client 
        
        response = client.chat.completions.create(
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            model="llama-3.3-70b-versatile",
            temperature=0.1 
        )
        
        return {"answer": response.choices[0].message.content}

    except Exception as e:
        print(f"❌ Global Search Error: {e}")
        return {"answer": "Sorry, I ran into an error while searching across your meetings."}